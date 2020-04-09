from django.shortcuts import render,get_object_or_404, redirect
from django.http import HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.core.exceptions import ObjectDoesNotExist
from django.views.generic import (
    ListView,
    DetailView,
    CreateView,
    DeleteView,
    UpdateView
)
from olahDataMatakuliah.models import matakuliah
from olahDataJurnalKuliah.models import jurnalKuliah,detilJurnalKuliah, pesertaKuliah
from olahDataRPS.models import rps, detilRPS
from .forms import detilJurnalKuliahForm, updatedetilJurnalKuliahForm
from presensiKuliah.models import presensi
from olahDataMahasiswa.models import mahasiswa
import openpyxl
# Create your views here.

# menampilkan RPS dengan kode mk yang sama pada jurnal kuliah

class detilRPSListView(ListView): #rubah menjadi detailView
    model = detilRPS
    template_name = 'subPokokBahasan/listRpsUntukJurnal.html'
    # paginate_by = 10
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Pilih Pertemuan untuk jurnal kuliah', 
    }

    def get_queryset(self):
        if not self.kwargs['id_rps'] == None:
            self.queryset=self.model.objects.filter(idRps_id=self.kwargs['id_rps'])
        return super().get_queryset()

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        print (kwargs)
        return super().get_context_data(*args, **kwargs)



class jurnalKuliahListView(ListView):
    model = jurnalKuliah
    template_name = 'subPokokBahasan/jurnalKuliah_list.html'
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Sub Pokok Bahasan - Pilih daftar jurnal kuliah', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        return super().get_context_data(*args, **kwargs)


class jurnalKuliahDetailView(DetailView):
    model = jurnalKuliah
    template_name = 'subPokokBahasan/jurnalKuliah_detail.html'
    extra_context ={
        'appGroup':'Presensi',
        'appName':'Sub Pokok Bahasan', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        idJurnal = jurnalKuliah.objects.values_list('mk_id', flat=True).get(id=self.kwargs['pk'])
        self.kwargs.update({'idJurnal':idJurnal})
        matkul = matakuliah.objects.all().get(id=self.kwargs['idJurnal'])
        self.kwargs.update({'matkul':matkul})

        detilJurnal=detilJurnalKuliah.objects.filter(jurnal_id=self.kwargs['pk'])
        self.kwargs.update({'detilJurnal':detilJurnal}) 

        # ambil id_rps yang memiliki kode matakuliah 
        # jika ada id_rps maka
        try:
            id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        # except id_rps.DoesNotExist:
        except ObjectDoesNotExist:
            id_rps = None
        # id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        self.kwargs.update({'id_rps':id_rps})

        kwargs = self.kwargs
        print(kwargs)
        print('----------')
        print(kwargs['matkul'])
        print('----------')
        print(self.object.mk_id)
        return super().get_context_data(*args, **kwargs)



class detilJurnalKuliahCreateView(CreateView):
    form_class = detilJurnalKuliahForm
    template_name = 'subPokokBahasan/createdetilJurnalKualiah.html'
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Tambah Pokok Bahasan', 
    }

    def get_context_data(self,*args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        print('dari get_context_data')
        print(kwargs)
        return super().get_context_data(*args, **kwargs)

    def get_initial(self, *args, **kwargs):
        jurnal = get_object_or_404(jurnalKuliah,id=self.kwargs['id_jurnal'])
        print('dari get initial')
        print (kwargs)
        return{
            'jurnal':jurnal
        }


class detilJurnalKuliahFromRPSCreateView(CreateView):
    form_class = detilJurnalKuliahForm
    template_name = 'subPokokBahasan/createdetilJurnalKualiah.html'
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Tambah Pokok Bahasan', 
    }

    def get_context_data(self,*args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        print('dari get_context_data')
        print(kwargs)
        return super().get_context_data(*args, **kwargs)

    def get_initial(self, *args, **kwargs):
        jurnal = get_object_or_404(jurnalKuliah,id=self.kwargs['id_jurnal'])

        # id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        # ambil isi pertemuan dari detilRPS berdasar id detilRPS
        pertemuan = detilRPS.objects.values_list('pertemuan', flat=True).get(id=self.kwargs['id_rps'])

        # ambil isi materiBelajar dari detilRPS berdasar id detilRPS
        materiBelajar = detilRPS.objects.values_list('materiBelajar', flat=True).get(id=self.kwargs['id_rps'])
        
        # ambil isi bentukMetodeBelajar dari detilRPS berdasar id detilRPS
        bentukMetodeBelajar = detilRPS.objects.values_list('bentukMetodeBelajar', flat=True).get(id=self.kwargs['id_rps'])

        print('dari get initial')
        print (kwargs)
        print('=============')
        print(jurnal)
        print(pertemuan)
        return{
            'jurnal':jurnal,
            'pertemuan':pertemuan,
            'materi':materiBelajar,
            'metode':bentukMetodeBelajar
        }


class detilJurnalKuliahUpdateView(UpdateView):
    model = detilJurnalKuliah
    form_class = updatedetilJurnalKuliahForm
    template_name = 'subPokokBahasan/createdetilJurnalKualiah.html'
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Tambah Pokok Bahasan', 
    }

    def get_context_data(self,*args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        print(kwargs)
        return super().get_context_data(*args, **kwargs)

    # def get_initial(self):
    #     jurnal = get_object_or_404(jurnalKuliah,id=self.kwargs['id_jurnal'])
    #     return{
    #         'jurnal':jurnal
    #     }



class detilJurnalKuliahDeleteView(DeleteView):
    model = detilJurnalKuliah
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Hapus Pokok Bahasan', 
    }

    def get_context_data(self,*args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        # print(kwargs)
        return super().get_context_data(*args, **kwargs)
    
    def get_success_url(self):
        jurnal = self.object.jurnal
        return reverse_lazy('subPokokBahasan:detiljurnalkuliah', kwargs={'pk':jurnal.id})


# UNTUK PRESENSI

class presensiDetailView(DetailView):
    model = jurnalKuliah
    template_name = 'subPokokBahasan/presensi_form.html'
    extra_context = {
        'appGroup':'Presensi',
        'appName':'Presensi kuliah', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)

        dataKuliah = detilJurnalKuliah.objects.all().get(id=self.kwargs['id_dtJurnal'])
        self.kwargs.update({'dataKuliah':dataKuliah})

        pesertaKuliah = presensi.objects.filter(
            jurnalPerkuliahan_id=self.kwargs['id_dtJurnal']
        ).order_by(
            '-presensi',
            '-presenceDate'
        ).values(
            'npm__npm',
            'npm__nama',
            'presensi',
            'presenceDate'
        )
        self.kwargs.update({'pesertaKuliah':pesertaKuliah})

        jumlahKehadiran = presensi.objects.filter(
            jurnalPerkuliahan_id=self.kwargs['id_dtJurnal'],
            presensi=True
        ).count()
        self.kwargs.update({'jumlahKehadiran':jumlahKehadiran})

        kwargs = self.kwargs
        # print(kwargs)
        # print(kwargs['jumlahKehadiran'])
        return super().get_context_data(*args, **kwargs)



def index(request):
    context = {
        'appGroup':'Presensi',
        'appName':'SuB Pokok Bahasan',
    }
    return render(request,'subPokokBahasan/index.html',context)



def importPesertaKuliah(request,id_dtjurnal):
    context = {
            'appGroup': 'Presensi',
            'appName': 'Import Mahasiswa Peserta Kuliah',
            'id_dtJurnal':id_dtjurnal
        }
    if request.method=='GET':
        return render (request,'subPokokBahasan/importPesertaKuliah.html',context)
    else:
        dataFile = request.FILES['fileImport']
        wb = openpyxl.load_workbook(dataFile)
        sheets = wb.sheetnames
        print(sheets)
        worksheet = wb[sheets[0]]

        exel_data=list()
        dataSudahAda = list()
        mhsBelumAda = list()
        # perulangan row 
        for row in worksheet.iter_rows(min_row=2, max_col=2):
            row_data = list()
            # perulangan kolom
            for cell in row:
                isi = str(cell.value)
                row_data.append(isi.replace("'",""))
            dataNpm = row_data[0]
            print(dataNpm)
            try:
                # ambil id mahasiswa
                idMhs = mahasiswa.objects.values_list('id', flat=True).get(npm=dataNpm)
                
                try:    
                    # cek id mahasiswa tersebut sudah ada di tabel presensi apa belum 
                    # dengan jurnalPerkuliahan_id=id_dtJurnal
                    presensi.objects.get(npm_id=idMhs,jurnalPerkuliahan_id=id_dtjurnal)
                    dataSudahAda.append(row_data)
                except ObjectDoesNotExist:
                    # jika belum tambahkan
                    presensi.objects.create(
                        jurnalPerkuliahan_id=id_dtjurnal,
                        npm_id=idMhs
                    )
                    exel_data.append(row_data)
                    # pass
            except ObjectDoesNotExist:
                # gagal diimport, karena data mahasiswa belum ditambahkan, 
                mhsBelumAda.append(row_data)
                # pass
        # asdda
        context['exel_data']=exel_data
        context['dataSudahAda']=dataSudahAda
        context['mhsBelumAda']=mhsBelumAda

        return render (request,'subPokokBahasan/importPesertaKuliah.html',context)

def buatPresensi(request, idJurnal,id_dtJurnal):
    context = {
            'appGroup': 'Presensi',
            'appName': 'Tambahkan Mahasiswa Peserta Kuliah',
        }
    # ambil mahasiswa pesertaKuliah
    mhsPeserta =  pesertaKuliah.objects.values_list(
            'peserta',
            flat=True
        ).filter(jurnal_id=idJurnal)

    for mhs in mhsPeserta:
        # ek di tabel presensi, sudah ada apa belum, jika sudah maka pass, 
        try:
            presensi.objects.get(
                jurnalPerkuliahan_id=id_dtJurnal,
                npm_id=mhs
            )
            # pass
        except ObjectDoesNotExist:
            # jika belum masukkan ke tabel presensi 
            presensi.objects.create(
                jurnalPerkuliahan_id=id_dtJurnal,
                npm_id=mhs
            )
        # print('ini mhs')
        # print(mhs)

    # context['mhsPeserta']=mhsPeserta
    # print('######################')
    # print(mhsPeserta)

    return HttpResponseRedirect(reverse('subPokokBahasan:presensi', kwargs={'pk':idJurnal,'id_dtJurnal':id_dtJurnal}))
