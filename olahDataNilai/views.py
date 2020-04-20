from django.shortcuts import render
from django.urls import reverse_lazy
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.views.generic import(
    CreateView, DetailView, ListView, 
    DeleteView, UpdateView
)
from olahDataJurnalKuliah.models import jurnalKuliah, detilJurnalKuliah, pesertaKuliah
from olahDataMatakuliah.models import matakuliah
from olahDataRPS.models import rps, detilRPS
from presensiKuliah.models import presensi
from .forms import updateNilaiPresensiForm
import openpyxl
from openpyxl import Workbook

# Create your views here.
def index(request):
    context = {
        'appGroup':'Dosen',
        'appName':'Olah Data Nilai',
    }
    return render(request,'olahDataNilai/index.html',context)

class jurnalKuliahListView(ListView):
    model = jurnalKuliah
    template_name = 'olahDataNilai/jurnalKuliah_list.html'
    extra_context = {
        'appGroup':'Dosen',
        'appName':'Olah Data Nilai. Silahkan pilih perkuliahan yang diampu', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        return super().get_context_data(*args, **kwargs)

class jurnalKuliahDetailView(DetailView):
    model = jurnalKuliah
    template_name = 'olahDataNilai/jurnalKuliah_detail.html'
    extra_context ={
        'appGroup':'Dosen',
        'appName':'Daftar Pokok Bahasan', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        idJurnal = jurnalKuliah.objects.values_list('mk_id', flat=True).get(id=self.kwargs['pk'])
        self.kwargs.update({'idJurnal':idJurnal})
        matkul = matakuliah.objects.all().get(id=self.kwargs['idJurnal'])
        self.kwargs.update({'matkul':matkul})

        detilJurnal=detilJurnalKuliah.objects.filter(jurnal_id=self.kwargs['pk'])
        self.kwargs.update({'detilJurnal':detilJurnal}) 

        # ambil id_rps yang memiliki kode matakuliah 
        # jika ada id_rps maka
        try:
            id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        # except id_rps.DoesNotExist:
        except ObjectDoesNotExist:
            id_rps = None
        # id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        self.kwargs.update({'id_rps':id_rps})

        kwargs = self.kwargs
        print(kwargs)
        print('----------')
        print(kwargs['matkul'])
        print('----------')
        print(self.object.mk_id)
        return super().get_context_data(*args, **kwargs)




class nilaiPerPertemuan(DetailView):
    model = jurnalKuliah
    template_name = 'olahDataNilai/nilaiPertemuan.html'
    extra_context = {
        'appGroup':'Dosen',
        'appName':'Nilai Harian', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)

        dataKuliah = detilJurnalKuliah.objects.all().get(id=self.kwargs['id_dtJurnal'])
        self.kwargs.update({'dataKuliah':dataKuliah})

        pesertaKuliah = presensi.objects.filter(
            jurnalPerkuliahan_id=self.kwargs['id_dtJurnal']
        ).order_by(
            '-presensi',
            '-presenceDate'
        ).values(
            'id',
            'npm__npm',
            'npm__nama',
            'presensi',
            'presenceDate',
            'nilai'
        )
        self.kwargs.update({'pesertaKuliah':pesertaKuliah})

        jumlahKehadiran = presensi.objects.filter(
            jurnalPerkuliahan_id=self.kwargs['id_dtJurnal'],
            presensi=True
        ).count()
        self.kwargs.update({'jumlahKehadiran':jumlahKehadiran})

        kwargs = self.kwargs
        print(kwargs)
        # print(kwargs['jumlahKehadiran'])
        return super().get_context_data(*args, **kwargs)



class nilaiUpdateView(UpdateView):
    model = presensi
    form_class = updateNilaiPresensiForm
    template_name = 'olahDataNilai/nilai.html'
    extra_context = {
        'appGroup':'Olah Data Nilai',
        'appName':'Nilai Harian', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)

        namaMhs = presensi.objects.values_list(
            'npm__nama', flat=True
        ).get(id=self.kwargs['pk'])
        self.kwargs.update({'namaMhs':namaMhs})

        kwargs = self.kwargs
        print('-----------------------')
        print(kwargs)
        return super().get_context_data(*args, **kwargs)
    
    def get_success_url(self):
        # jurnalKuliah = self.object.jurnalKuliah
        jurnalPerkuliahan = self.object.jurnalPerkuliahan
        return reverse_lazy('olahDataNilai:nilaiperpertemuan', kwargs={
            "pk": jurnalPerkuliahan.jurnal_id,
            "id_dtJurnal":jurnalPerkuliahan.id
            })

def exportNilaiHarian(request, jurnalPerkuliahan):
    context = {
            'appGroup':'Olah Data Nilai',
            'appName':'Export/Import Nilai per pertemuan',
            'jurnalPerkuliahan':jurnalPerkuliahan
        }
    if request.method=='GET':
        # Download file excel per pertemuan
        dataMhs = presensi.objects.filter(
            jurnalPerkuliahan_id=jurnalPerkuliahan
            ).values(
                'id',
                'npm__npm',
                'npm__nama',
                'presensi',
                'nilai'
            )
        # pertemuan = 
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response['Content-Disposition']='attachment; filename=dataNilaiHarian.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.tittle='NilaiHarian'

        kolom = [
            'ID',
            'NPM',
            'Nama',
            'Hadir',
            'Nilai'
        ]

        row_num = 1

        for col_num, column_tittle in enumerate(kolom, 1):
            cell=ws.cell(row = row_num,column = col_num)
            cell.value = column_tittle

        for mhs in dataMhs:
            row_num+=1
            # Definisi data untuk setiap baris 
            row = [
                mhs['id'],
                mhs['npm__npm'],
                mhs['npm__nama'],
                mhs['presensi'],
                mhs['nilai']
            ]
            # masukkan data ke cell
            for col_num, cell_value in enumerate(row,1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value=cell_value
        
        wb.save(response)
        return response

def importNilaiHarian(request,idJurnal,id_dtJurnal):
    context = {
            'appGroup':'Olah Data Nilai',
            'appName':'Import Nilai Harian',
            'idJurnal':idJurnal,
            'id_dtJurnal':id_dtJurnal
        }
    print(context)
    
    if request.method=='GET':
        return render (request,'olahDataNilai/importNilaiHarian.html',context)
    
    else:
        dataNilaiHarian = request.FILES['fileImport']
        wb=openpyxl.load_workbook(dataNilaiHarian)
        sheets = wb.sheetnames
        ws=wb[sheets[0]]

        exel_data=list()
        for row in ws.iter_rows(min_row=2, max_col=5):
            row_data=list()
            for cell in row:
                row_data.append(cell.value)

            exel_data.append(row_data)

            idPresensi = row_data[0]
            nilaiHarian = row_data[4]
            # perintah update di sini
            presensi.objects.filter(id=idPresensi).update(nilai=nilaiHarian)

            # print('id : '+ str(idPresensi))
            # print('nilai : '+ str(nilaiHarian))

        context['exel_data']=exel_data

        return render(request,'olahDataNilai/importNilaiHarian.html',context)


class rekapTotalDetailView(DetailView):
    model = jurnalKuliah
    template_name = 'olahDataNilai/rekapTotal.html'
    extra_context ={
        'appGroup':'Dosen',
        'appName':'Rekap Nilai Akhir Matakuliah', 
    }  

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        idJurnal = jurnalKuliah.objects.values_list('mk_id', flat=True).get(id=self.kwargs['pk']) #mencari kode_mk
        self.kwargs.update({'idJurnal':idJurnal})
        matkul = matakuliah.objects.all().get(id=self.kwargs['idJurnal'])
        self.kwargs.update({'matkul':matkul})

        detilJurnal=detilJurnalKuliah.objects.filter(jurnal_id=self.kwargs['pk'])
        self.kwargs.update({'detilJurnal':detilJurnal}) 

        jmlPertemuan = detilJurnal.count()
        self.kwargs.update({'jmlPertemuan':jmlPertemuan})

        # ambil id_rps yang memiliki kode matakuliah 
        # jika ada id_rps maka
        try:
            id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        # except id_rps.DoesNotExist:
        except ObjectDoesNotExist:
            id_rps = None
        # id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=self.object.mk_id)
        self.kwargs.update({'id_rps':id_rps})



        # rekap daftar nilai ke dalam list
        daftarMhs = pesertaKuliah.objects.filter(
            jurnal_id=self.kwargs['pk']
        ).values(
            'peserta__npm',
            'peserta__nama'
        ).order_by('-peserta__npm')
        baris = list()
        for mhs in daftarMhs:
            dataBaris = list()
            dataBaris.append(mhs['peserta__npm'])
            dataBaris.append(mhs['peserta__nama'])
            # ambil pertemuan, loop pertemuan pada jurnal tersebut
            
            nilaixBobot = list() #list untuk menampung nilai x bobot
            for pert in detilJurnal:
                # cari nilai setiap pertemuan
                nilaiMhs=presensi.objects.values(
                    'nilai'
                ).get(
                    npm__npm=mhs['peserta__npm'],
                    jurnalPerkuliahan__pertemuan=pert.pertemuan
                )

                dataBaris.append(nilaiMhs['nilai'])
                
                # dan ambil bobot nilai per pertemuan dalam detil RPS berdasar id_rps
                bobotNilai=detilRPS.objects.values(
                    'bobotPenilaian'
                ).get(
                    idRps_id = id_rps,
                    pertemuan = pert.pertemuan
                )

                nilaiTerbobot = (bobotNilai['bobotPenilaian']/100)*nilaiMhs['nilai']
                nilaixBobot.append(nilaiTerbobot)
            
            nilaiAkhir = sum(nilaixBobot)
            dataBaris.append(nilaiAkhir)

            # tampung ke dalam satu baris
            baris.append(dataBaris)

        # masukkan ke dalam extra_context
        self.extra_context['baris']=baris 
        self.kwargs.update(self.extra_context) 



        # peserta kuliah
        pstKuliah = pesertaKuliah.objects.filter(
            jurnal_id=self.kwargs['pk']
        ).values(
            'peserta__npm',
            'peserta__nama'
        )
        self.kwargs.update({'pstKuliah':pstKuliah})     

        kwargs = self.kwargs
        # print(kwargs)
        # print('----------')
        # print(kwargs['matkul'])
        # print('----------')
        # print(self.object.mk_id)
        return super().get_context_data(*args, **kwargs)

def exportNilaiAkhir(request,idJurnal):
    idMk = jurnalKuliah.objects.values_list('mk_id', flat=True).get(id=idJurnal)
    matkul = matakuliah.objects.values('nama').get(id=idMk)
    detilJurnal=detilJurnalKuliah.objects.filter(jurnal_id=idJurnal)
    kelas = jurnalKuliah.objects.values_list('kelas',flat=True).get(id=idJurnal)
    # ambil id_rps yang memiliki kode matakuliah 
    # jika ada id_rps maka
    try:
        id_rps=rps.objects.values_list('id',flat=True).get(kodemk_id=idMk)
    # except id_rps.DoesNotExist:
    except ObjectDoesNotExist:
        id_rps = None
    

    # rekap daftar nilai ke dalam list
    daftarMhs = pesertaKuliah.objects.filter(
        jurnal_id=idJurnal
    ).values(
        'peserta__npm',
        'peserta__nama'
    ).order_by('-peserta__npm')
    baris = list()
    for mhs in daftarMhs:
        dataBaris = list()
        dataBaris.append(mhs['peserta__npm'])
        dataBaris.append(mhs['peserta__nama'])
        # ambil pertemuan, loop pertemuan pada jurnal tersebut
        
        nilaixBobot = list() #list untuk menampung nilai x bobot
        for pert in detilJurnal:
            # cari nilai setiap pertemuan
            nilaiMhs=presensi.objects.values(
                'nilai'
            ).get(
                npm__npm=mhs['peserta__npm'],
                jurnalPerkuliahan__pertemuan=pert.pertemuan
            )

            dataBaris.append(nilaiMhs['nilai'])
            
            # dan ambil bobot nilai per pertemuan dalam detil RPS berdasar id_rps
            bobotNilai=detilRPS.objects.values(
                'bobotPenilaian'
            ).get(
                idRps_id = id_rps,
                pertemuan = pert.pertemuan
            )

            nilaiTerbobot = (bobotNilai['bobotPenilaian']/100)*nilaiMhs['nilai']
            nilaixBobot.append(nilaiTerbobot)
        
        nilaiAkhir = sum(nilaixBobot)
        dataBaris.append(nilaiAkhir)

        # tampung ke dalam satu baris
        baris.append(dataBaris)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition']='attachment; filename=NilaiTotal-{mk}-{kls}.xlsx'.format(mk=matkul['nama'],kls=kelas)
    wb = Workbook()
    ws = wb.active
    ws.tittle='NilaiTotal'
    kolom = [
        'NPM',
        'NAMA',
    ]

    # menambak kolom nilai pada header
    for itemPertemuan in detilJurnal:
        n = itemPertemuan.pertemuan
        kolom.append('N'+str(n))

    kolom.append('Nilai Akhir')

    row_num = 1

    # membuat header excel
    for col_num, column_tittle in enumerate(kolom,1):
        cell = ws.cell(row = row_num, column = col_num)
        cell.value = column_tittle

    for dataNilai in baris:
        ws.append(dataNilai)
        
    wb.save(response)

    return response