from django.shortcuts import render
from django.db import transaction
from django.urls import reverse_lazy
from django.core.exceptions import ObjectDoesNotExist
from django.views.generic import CreateView, DeleteView, ListView, UpdateView, DetailView
from . forms import (
        jurnalKuliahForm, 
        detilJurnalKuliahForm, 
        JurnalFormset, 
        jurnalKuliahForm
    )
from . models import jurnalKuliah, detilJurnalKuliah, pesertaKuliah
from olahDataMahasiswa.models import mahasiswa
import openpyxl


# Create your views here.
def index (request):
    context = {
        'appGroup' : 'Operasional',
        'appName' : 'Olah Data Jurnal Kuliah',
    }
    return render(request,'olahDataJurnalKuliah/index.html',context)


class jurnalKuliahListView(ListView):
    model = jurnalKuliah
    ordering =['-id']
    extra_context = {
        'appGroup' : 'Operasional',
        'appName' : 'Olah Data Jurnal Kuliah',
    }
    
    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        return super().get_context_data(*args, **kwargs)


class jurnalKuliahCreateView(CreateView):
    form_class = jurnalKuliahForm
    template_name = 'olahDataJurnalKuliah/jurnalkuliah_form.html'
    extra_context = {
        'appGroup':'Operasional',
        'appName':'Olah Data Jurnal Kuliah', 
    }

    def get_context_data(self,*args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        # print(kwargs)
        return super().get_context_data(*args, **kwargs)
    # model = jurnalKuliah

    # fields = [
    #     'mk',
    #     'tahunAjaran',
    #     'semester',
    #     'dosenPengajar',
    #     'ruang',
    #     'pjmk'
    # ]


class jurnalPerkuliahanCreateView(CreateView):
    model = jurnalKuliah
    fields = [
        'mk',
        'tahunAjaran',
        'semester',
        'dosenPengajar',
        'ruang',
        'pjmk'
    ]
    success_url = reverse_lazy('olahDataJurnalKuliah:jurnalKuliahList')

    def get_context_data(self, **kwargs):
        data = super(jurnalPerkuliahanCreateView, self).get_context_data(**kwargs)
        if self.request.POST :
            data['dtlJurnalKuliah']=JurnalFormset(self.request.POST)
        else:
            data['dtlJurnalKuliah']=JurnalFormset()
        return data
    
    def form_valid(self, form):
        context = self.get_context_data()
        dtlJurnalKuliah = context['dtlJurnalKuliah']
        with transaction.atomic():
            self.object = form.save()
            if dtlJurnalKuliah.is_valid():
                dtlJurnalKuliah.instance=self.object
                dtlJurnalKuliah.save()
        return super(jurnalPerkuliahanCreateView,self).form_valid(form)


class jurnalKuliahUpdateView(UpdateView):
    model = jurnalKuliah
    success_url = reverse_lazy('olahDataJurnalKuliah:jurnalKuliahList')
    fields = [
        'mk',
        'tahunAjaran',
        'semester',
        'dosenPengajar',
        'ruang',
        'pjmk'
    ]


class jurnalPerkuliahanUpdateView(UpdateView):
    form_class = jurnalKuliahForm
    model = jurnalKuliah
    template_name = 'olahDataJurnalKuliah/jurnalkuliah_form.html'
    extra_context = {
        'appGroup':'Operasional',
        'appName':'Olah Data Jurnal Kuliah', 
    }
    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)
        kwargs = self.kwargs
        return super().get_context_data(*args, **kwargs)

    # fields = [
    #     'mk',
    #     'tahunAjaran',
    #     'semester',
    #     'dosenPengajar',
    #     'ruang',
    #     'pjmk'
    # ]
    success_url = reverse_lazy('olahDataJurnalKuliah:jurnalKuliahList')
    
    # def get_context_data(self,**kwargs):
    #     data = super(jurnalPerkuliahanUpdateView, self).get_context_data(**kwargs)
    #     if self.request.POST :
    #         data['dtlJurnalKuliah']=JurnalFormset(self.request.POST, instance=self.object)
    #     else:
    #         data['dtlJurnalKuliah']=JurnalFormset(instance = self.object)
    #     return data
    
    # def form_valid(self, form):
    #     context = self.get_context_data()
    #     dtlJurnalKuliah = context['dtlJurnalKuliah']
    #     with transaction.atomic():
    #         self.object = form.save()
    #         if dtlJurnalKuliah.is_valid():
    #             dtlJurnalKuliah.instance=self.object
    #             dtlJurnalKuliah.save()
    #     return super(jurnalPerkuliahanUpdateView,self).form_valid(form)
    

class jurnalKuliahDeleteView(DeleteView):
    model = jurnalKuliah
    success_url = reverse_lazy('olahDataJurnalKuliah:jurnalKuliahList')


class jurnalKuliahDetailView(DetailView):
    model = jurnalKuliah
    template_name = 'olahDataJurnalKuliah/jurnalKuliah_detail.html'
    extra_context = {
        'appGroup':'Operasional',
        'appName':'Olah Data Jurnal Kuliah', 
    }

    def get_context_data(self, *args, **kwargs):
        self.kwargs.update(self.extra_context)

        # tampilkan data peserta kuliah
        pstKuliah = pesertaKuliah.objects.filter(
            jurnal_id=self.kwargs['pk']
            ).values(
                'peserta__npm',
                'peserta__nama'
            )
        self.kwargs.update({'pstKuliah':pstKuliah})

        kwargs = self.kwargs
        print(self.object)
        return super().get_context_data(*args, **kwargs)


def importPesertaKuliah(request,idJurnal):
    context = {
            'appGroup': 'Presensi',
            'appName': 'Import Mahasiswa Peserta Kuliah',
            'idJurnal':idJurnal
        }
    if request.method=='GET':
        return render (request,'olahDataJurnalKuliah/importPesertaKuliah.html',context)
    else:
        dataFile = request.FILES['fileImport']
        wb = openpyxl.load_workbook(dataFile)
        sheets = wb.sheetnames
        print(sheets)
        worksheet = wb[sheets[0]]

        exel_data=list()
        dataSudahAda = list()
        mhsBelumAda = list()
        # perulangan row 
        for row in worksheet.iter_rows(min_row=2, max_col=2):
            row_data = list()
            # perulangan kolom
            for cell in row:
                isi = str(cell.value)
                row_data.append(isi.replace("'",""))
            dataNpm = row_data[0]
            print(dataNpm)
            try:
                # ambil id mahasiswa
                idMhs = mahasiswa.objects.values_list('id', flat=True).get(npm=dataNpm)
                
                try:    
                    # cek id mahasiswa tersebut sudah ada di tabel pesertaKuliah apa belum 
                    # dengan idJurnalKuliah=idJurnal
                    pesertaKuliah.objects.get(peserta_id=idMhs,jurnal_id=idJurnal)
                    dataSudahAda.append(row_data)
                except ObjectDoesNotExist:
                    # jika belum tambahkan
                    pesertaKuliah.objects.create(
                        jurnal_id=idJurnal,
                        peserta_id=idMhs
                    )
                    exel_data.append(row_data)
                    # pass
            except ObjectDoesNotExist:
                # gagal diimport, karena data mahasiswa belum ditambahkan, 
                mhsBelumAda.append(row_data)
                # pass
        # asdda
        context['exel_data']=exel_data
        context['dataSudahAda']=dataSudahAda
        context['mhsBelumAda']=mhsBelumAda

        return render (request,'olahDataJurnalKuliah/importPesertaKuliah.html',context)